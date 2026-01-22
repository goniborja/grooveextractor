"""Exportador de datos de groove a Excel."""
from typing import Dict, List, Optional
from pathlib import Path
import hashlib
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..models import (
    GrooveData, InstrumentData, GridMapping,
    HumanizationStats, SwingAnalysis, JamaicanStyle,
    STYLE_BPM_RANGES
)


# Catalogo de instrumentos MIDI (GM Drums)
INSTRUMENT_CATALOG = {
    "kick": {"nota_midi": 36, "default_vel": 100, "descripcion": "Bass Drum 1"},
    "snare": {"nota_midi": 38, "default_vel": 100, "descripcion": "Acoustic Snare"},
    "rimshot": {"nota_midi": 37, "default_vel": 90, "descripcion": "Side Stick"},
    "hihat": {"nota_midi": 42, "default_vel": 80, "descripcion": "Closed Hi-Hat"},
    "hihat_closed": {"nota_midi": 42, "default_vel": 80, "descripcion": "Closed Hi-Hat"},
    "hihat_open": {"nota_midi": 46, "default_vel": 85, "descripcion": "Open Hi-Hat"},
    "hihat_pedal": {"nota_midi": 44, "default_vel": 70, "descripcion": "Pedal Hi-Hat"},
    "tom_high": {"nota_midi": 50, "default_vel": 95, "descripcion": "High Tom"},
    "tom_mid": {"nota_midi": 47, "default_vel": 95, "descripcion": "Low-Mid Tom"},
    "tom_low": {"nota_midi": 45, "default_vel": 95, "descripcion": "Low Tom"},
    "crash": {"nota_midi": 49, "default_vel": 100, "descripcion": "Crash Cymbal 1"},
    "ride": {"nota_midi": 51, "default_vel": 85, "descripcion": "Ride Cymbal 1"},
}


class ExcelExporter:
    """
    Exporta datos de groove a formato Excel.

    Genera hojas:
    - INFO: Informacion general del analisis
    - REJILLAS: Patrones de grid (0/1) por instrumento con metadatos
    - HUMANIZACION: Velocidades y desviaciones de timing

    Formato compatible con DAWs y herramientas de produccion.
    """

    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl es requerido para exportar a Excel. "
                "Instalar con: pip install openpyxl"
            )

        # Estilos
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, size=11, color="FFFFFF")
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, groove_data: GrooveData, output_path: str) -> str:
        """
        Exporta GrooveData a archivo Excel.

        Args:
            groove_data: Datos del groove a exportar
            output_path: Ruta del archivo de salida (.xlsx)

        Returns:
            Ruta del archivo creado
        """
        wb = openpyxl.Workbook()

        # Eliminar hoja por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Crear hojas
        self._create_info_sheet(wb, groove_data)
        self._create_rejillas_sheet(wb, groove_data)
        self._create_humanizacion_sheet(wb, groove_data)

        # Guardar
        output_path = Path(output_path)
        if not output_path.suffix:
            output_path = output_path.with_suffix('.xlsx')

        wb.save(str(output_path))
        return str(output_path)

    def _generate_groove_id(self, song_name: str, inst_name: str, bar_idx: int) -> str:
        """Genera un ID unico para un groove."""
        data = f"{song_name}_{inst_name}_{bar_idx}"
        return hashlib.md5(data.encode()).hexdigest()[:8]

    def _calculate_timing_variance(self, timing_deviations: List[float]) -> float:
        """Calcula la variacion promedio de timing."""
        non_zero = [abs(t) for t in timing_deviations if t != 0]
        return round(sum(non_zero) / len(non_zero), 2) if non_zero else 0.0

    def _calculate_velocity_variance(self, velocities: List[int]) -> float:
        """Calcula la variacion de velocity."""
        non_zero = [v for v in velocities if v > 0]
        if len(non_zero) < 2:
            return 0.0
        mean = sum(non_zero) / len(non_zero)
        variance = sum((v - mean) ** 2 for v in non_zero) / len(non_zero)
        return round(variance ** 0.5, 2)  # Desviacion estandar

    def _create_info_sheet(self, wb, groove_data: GrooveData):
        """Crea hoja de informacion general."""
        ws = wb.create_sheet("INFO")

        # Encabezado
        ws['A1'] = "GROOVE EXTRACTOR - Analisis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Datos generales
        info_data = [
            ("Cancion", groove_data.song_name),
            ("BPM", f"{groove_data.bpm:.1f}"),
            ("Estilo", groove_data.style.value),
            ("Vintage", "Si" if groove_data.is_vintage else "No"),
            ("Tempo Drift", f"{groove_data.tempo_drift * 100:.2f}%"),
        ]

        if groove_data.swing:
            info_data.extend([
                ("Swing %", f"{groove_data.swing.swing_percentage:.1f}%"),
                ("Swing Ratio", f"{groove_data.swing.swing_ratio:.2f}"),
                ("Descripcion Swing", groove_data.swing.description),
            ])

        row = 3
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.header_font
            ws[f'B{row}'] = value
            row += 1

        # Instrumentos
        row += 1
        ws[f'A{row}'] = "Instrumentos Analizados"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for name, inst in groove_data.instruments.items():
            ws[f'A{row}'] = name
            ws[f'B{row}'] = f"{len(inst.onsets)} onsets, {len(inst.grids)} compases"
            row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _create_rejillas_sheet(self, wb, groove_data: GrooveData):
        """
        Crea hoja REJILLAS con patrones de grid y metadatos extendidos.

        Formato extendido:
        GROOVE_ID | ID_PATRON | INSTRUMENTO | ESTILO | BPM | SWING_AMOUNT |
        1-16 | VEL_BASE | TIMING_VAR | VEL_VAR | FUENTE | NOTA_MIDI
        """
        ws = wb.create_sheet("REJILLAS")

        # Encabezados extendidos
        headers = [
            "GROOVE_ID", "ID_PATRON", "INSTRUMENTO", "ESTILO", "BPM", "SWING_AMOUNT"
        ]
        headers += [str(i) for i in range(1, 17)]  # Pasos 1-16
        headers += ["VEL_BASE", "TIMING_VAR", "VEL_VAR", "FUENTE", "NOTA_MIDI"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Datos
        row = 2
        swing_amount = groove_data.swing.swing_percentage if groove_data.swing else 50.0

        for inst_name, inst_data in groove_data.instruments.items():
            nota_midi = INSTRUMENT_CATALOG.get(inst_name, {}).get("nota_midi", 0)

            for bar_idx, grid in enumerate(inst_data.grids):
                col = 1

                # GROOVE_ID
                groove_id = self._generate_groove_id(groove_data.song_name, inst_name, bar_idx)
                ws.cell(row=row, column=col, value=groove_id)
                col += 1

                # ID_PATRON
                ws.cell(row=row, column=col, value=f"{inst_name}_{bar_idx + 1}")
                col += 1

                # INSTRUMENTO
                ws.cell(row=row, column=col, value=inst_name)
                col += 1

                # ESTILO
                ws.cell(row=row, column=col, value=groove_data.style.value)
                col += 1

                # BPM
                ws.cell(row=row, column=col, value=round(groove_data.bpm, 1))
                col += 1

                # SWING_AMOUNT
                ws.cell(row=row, column=col, value=round(swing_amount, 1))
                col += 1

                # Columnas 1-16 (pattern)
                for step_idx, value in enumerate(grid.pattern):
                    cell = ws.cell(row=row, column=col + step_idx, value=value)
                    cell.alignment = self.center_align
                    if value == 1:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                col += 16

                # VEL_BASE (promedio de velocidades no-cero)
                non_zero_vels = [v for v in grid.velocities if v > 0]
                vel_base = int(sum(non_zero_vels) / len(non_zero_vels)) if non_zero_vels else 0
                ws.cell(row=row, column=col, value=vel_base)
                col += 1

                # TIMING_VAR
                timing_var = self._calculate_timing_variance(grid.timing_deviations)
                ws.cell(row=row, column=col, value=timing_var)
                col += 1

                # VEL_VAR
                vel_var = self._calculate_velocity_variance(grid.velocities)
                ws.cell(row=row, column=col, value=vel_var)
                col += 1

                # FUENTE
                ws.cell(row=row, column=col, value=groove_data.song_name)
                col += 1

                # NOTA_MIDI
                ws.cell(row=row, column=col, value=nota_midi)

                row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 12  # GROOVE_ID
        ws.column_dimensions['B'].width = 15  # ID_PATRON
        ws.column_dimensions['C'].width = 12  # INSTRUMENTO
        ws.column_dimensions['D'].width = 12  # ESTILO
        ws.column_dimensions['E'].width = 6   # BPM
        ws.column_dimensions['F'].width = 12  # SWING_AMOUNT
        for col in range(7, 23):  # Pasos 1-16
            ws.column_dimensions[get_column_letter(col)].width = 4
        ws.column_dimensions['W'].width = 10  # VEL_BASE
        ws.column_dimensions['X'].width = 10  # TIMING_VAR
        ws.column_dimensions['Y'].width = 10  # VEL_VAR
        ws.column_dimensions['Z'].width = 20  # FUENTE
        ws.column_dimensions['AA'].width = 10  # NOTA_MIDI

    def _create_humanizacion_sheet(self, wb, groove_data: GrooveData):
        """
        Crea hoja HUMANIZACION con velocidades y timing.

        Formato:
        GROOVE_ID | ID_PATRON | INSTRUMENTO | V1-V16 | T1-T16 | DURATION | FUENTE
        """
        ws = wb.create_sheet("HUMANIZACION")

        # Encabezados
        headers = ["GROOVE_ID", "ID_PATRON", "INSTRUMENTO"]
        headers += [f"V{i}" for i in range(1, 17)]  # Velocidades
        headers += [f"T{i}" for i in range(1, 17)]  # Timing deviations
        headers += ["DURATION", "FUENTE"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Datos
        row = 2

        for inst_name, inst_data in groove_data.instruments.items():
            for bar_idx, grid in enumerate(inst_data.grids):
                col = 1

                # GROOVE_ID
                groove_id = self._generate_groove_id(groove_data.song_name, inst_name, bar_idx)
                ws.cell(row=row, column=col, value=groove_id)
                col += 1

                # ID_PATRON
                ws.cell(row=row, column=col, value=f"{inst_name}_{bar_idx + 1}")
                col += 1

                # INSTRUMENTO
                ws.cell(row=row, column=col, value=inst_name)
                col += 1

                # V1-V16 (velocidades)
                for vel in grid.velocities:
                    cell = ws.cell(row=row, column=col, value=vel)
                    cell.alignment = self.center_align
                    col += 1

                # T1-T16 (timing deviations en ms)
                for dev in grid.timing_deviations:
                    cell = ws.cell(row=row, column=col, value=round(dev, 2))
                    cell.alignment = self.center_align
                    if dev < -5:  # Rushing
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    elif dev > 5:  # Dragging
                        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    col += 1

                # DURATION (duracion del compas en segundos)
                bar_duration = 4 * (60.0 / groove_data.bpm)  # 4 beats
                ws.cell(row=row, column=col, value=round(bar_duration, 3))
                col += 1

                # FUENTE
                ws.cell(row=row, column=col, value=groove_data.song_name)

                row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        for col in range(4, 36):
            ws.column_dimensions[get_column_letter(col)].width = 5
        ws.column_dimensions[get_column_letter(36)].width = 10
        ws.column_dimensions[get_column_letter(37)].width = 20

    def export_simple(self, patterns: Dict[str, List[List[int]]],
                      velocities: Dict[str, List[List[int]]],
                      bpm: float,
                      output_path: str) -> str:
        """
        Exportacion simplificada sin GrooveData completo.

        Args:
            patterns: Dict de instrumento -> lista de patrones (cada uno 16 valores 0/1)
            velocities: Dict de instrumento -> lista de velocidades (cada uno 16 valores 0-127)
            bpm: Tempo en BPM
            output_path: Ruta de salida

        Returns:
            Ruta del archivo creado
        """
        from ..models import OnsetList

        groove = GrooveData(song_name="export", bpm=bpm)

        for inst_name in patterns:
            inst_patterns = patterns[inst_name]
            inst_velocities = velocities.get(inst_name, [[100]*16] * len(inst_patterns))

            onsets = OnsetList(onsets=[], instrument=inst_name)
            grids = []

            for i, (pattern, vels) in enumerate(zip(inst_patterns, inst_velocities)):
                grid = GridMapping(
                    pattern=pattern,
                    velocities=vels,
                    timing_deviations=[0.0] * 16
                )
                grids.append(grid)

            inst_data = InstrumentData(name=inst_name, onsets=onsets, grids=grids)
            groove.add_instrument(inst_data)

        return self.export(groove, output_path)


class DatabaseAggregator:
    """
    Agrega datos de multiples analisis a una base de datos central (database.xlsx).

    La base de datos contiene:
    - ESTILOS: Catalogo de estilos jamaicanos
    - PATRONES: Metadatos de patrones analizados
    - REJILLAS: Patrones en grid de 16 pasos
    - HUMANIZACION: Datos de velocity y timing
    - INSTRUMENTOS: Catalogo de instrumentos MIDI
    """

    def __init__(self, database_path: str = "database.xlsx"):
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl es requerido. Instalar con: pip install openpyxl"
            )

        self.database_path = Path(database_path)

        # Estilos
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, size=11, color="FFFFFF")
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _create_base_database(self) -> openpyxl.Workbook:
        """Crea la estructura base de database.xlsx."""
        wb = openpyxl.Workbook()

        # Eliminar hoja por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Crear hojas con estructura
        self._create_estilos_sheet(wb)
        self._create_patrones_sheet(wb)
        self._create_rejillas_sheet(wb)
        self._create_humanizacion_sheet(wb)
        self._create_instrumentos_sheet(wb)

        return wb

    def _create_estilos_sheet(self, wb):
        """Crea hoja ESTILOS con catalogo de estilos."""
        ws = wb.create_sheet("ESTILOS")

        headers = ["ID", "NOMBRE", "BPM_MIN", "BPM_MAX", "FEEL", "SWING_TIPICO", "DESCRIPCION"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Datos de estilos jamaicanos
        estilos_data = [
            (1, "ska", 110, 140, "on-top", 55, "Ska clasico de los 60s, afterbeat"),
            (2, "rocksteady", 70, 90, "laid-back", 58, "Rocksteady, mas lento que ska"),
            (3, "early_reggae", 80, 100, "on-beat", 55, "Reggae temprano, transicion"),
            (4, "one_drop", 65, 85, "laid-back", 62, "One-drop clasico, bombo en 3"),
            (5, "rockers", 70, 90, "on-beat", 55, "Rockers, mas sincopado"),
            (6, "steppers", 70, 90, "on-beat", 52, "Steppers, bombo en todos los beats"),
            (7, "dub", 60, 80, "deep-pocket", 60, "Dub, espaciado y profundo"),
            (8, "dancehall", 85, 110, "on-top", 50, "Dancehall, mas moderno"),
        ]

        for row_idx, data in enumerate(estilos_data, 2):
            for col_idx, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Ajustar anchos
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 40

    def _create_patrones_sheet(self, wb):
        """Crea hoja PATRONES con metadatos."""
        ws = wb.create_sheet("PATRONES")

        headers = [
            "GROOVE_ID", "FUENTE", "ESTILO", "BPM", "SWING_AMOUNT",
            "FECHA_ANALISIS", "NUM_COMPASES", "DESCRIPCION"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 40

    def _create_rejillas_sheet(self, wb):
        """Crea hoja REJILLAS con estructura extendida."""
        ws = wb.create_sheet("REJILLAS")

        headers = [
            "GROOVE_ID", "ID_PATRON", "INSTRUMENTO", "ESTILO", "BPM", "SWING_AMOUNT"
        ]
        headers += [str(i) for i in range(1, 17)]
        headers += ["VEL_BASE", "TIMING_VAR", "VEL_VAR", "FUENTE", "NOTA_MIDI"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

    def _create_humanizacion_sheet(self, wb):
        """Crea hoja HUMANIZACION."""
        ws = wb.create_sheet("HUMANIZACION")

        headers = ["GROOVE_ID", "ID_PATRON", "INSTRUMENTO"]
        headers += [f"V{i}" for i in range(1, 17)]
        headers += [f"T{i}" for i in range(1, 17)]
        headers += ["DURATION", "FUENTE"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

    def _create_instrumentos_sheet(self, wb):
        """Crea hoja INSTRUMENTOS con catalogo."""
        ws = wb.create_sheet("INSTRUMENTOS")

        headers = ["ID", "NOMBRE", "NOTA_MIDI", "DEFAULT_VEL", "DESCRIPCION"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Datos del catalogo
        for idx, (name, info) in enumerate(INSTRUMENT_CATALOG.items(), 1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=name)
            ws.cell(row=idx + 1, column=3, value=info["nota_midi"])
            ws.cell(row=idx + 1, column=4, value=info["default_vel"])
            ws.cell(row=idx + 1, column=5, value=info["descripcion"])

        # Ajustar anchos
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25

    def _get_existing_groove_ids(self, ws) -> set:
        """Obtiene los GROOVE_IDs existentes en una hoja."""
        existing = set()
        for row in range(2, ws.max_row + 1):
            groove_id = ws.cell(row=row, column=1).value
            if groove_id:
                existing.add(groove_id)
        return existing

    def _generate_groove_id(self, song_name: str, inst_name: str, bar_idx: int) -> str:
        """Genera un ID unico para un groove."""
        data = f"{song_name}_{inst_name}_{bar_idx}"
        return hashlib.md5(data.encode()).hexdigest()[:8]

    def _calculate_timing_variance(self, timing_deviations: List[float]) -> float:
        """Calcula la variacion promedio de timing."""
        non_zero = [abs(t) for t in timing_deviations if t != 0]
        return round(sum(non_zero) / len(non_zero), 2) if non_zero else 0.0

    def _calculate_velocity_variance(self, velocities: List[int]) -> float:
        """Calcula la variacion de velocity."""
        non_zero = [v for v in velocities if v > 0]
        if len(non_zero) < 2:
            return 0.0
        mean = sum(non_zero) / len(non_zero)
        variance = sum((v - mean) ** 2 for v in non_zero) / len(non_zero)
        return round(variance ** 0.5, 2)

    def add_groove_data(self, groove_data: GrooveData) -> dict:
        """
        Agrega datos de un GrooveData a database.xlsx.

        Args:
            groove_data: Datos del groove a agregar

        Returns:
            Dict con estadisticas de la operacion
        """
        # Cargar o crear database
        if self.database_path.exists():
            wb = openpyxl.load_workbook(str(self.database_path))
        else:
            wb = self._create_base_database()

        # Obtener hojas
        ws_patrones = wb["PATRONES"]
        ws_rejillas = wb["REJILLAS"]
        ws_humanizacion = wb["HUMANIZACION"]

        # Obtener IDs existentes para evitar duplicados
        existing_rejillas = self._get_existing_groove_ids(ws_rejillas)
        existing_humanizacion = self._get_existing_groove_ids(ws_humanizacion)

        stats = {"added": 0, "skipped": 0, "song": groove_data.song_name}

        # Agregar metadatos a PATRONES
        patron_row = ws_patrones.max_row + 1
        song_groove_id = hashlib.md5(groove_data.song_name.encode()).hexdigest()[:8]
        swing_amount = groove_data.swing.swing_percentage if groove_data.swing else 50.0

        ws_patrones.cell(row=patron_row, column=1, value=song_groove_id)
        ws_patrones.cell(row=patron_row, column=2, value=groove_data.song_name)
        ws_patrones.cell(row=patron_row, column=3, value=groove_data.style.value)
        ws_patrones.cell(row=patron_row, column=4, value=round(groove_data.bpm, 1))
        ws_patrones.cell(row=patron_row, column=5, value=round(swing_amount, 1))
        ws_patrones.cell(row=patron_row, column=6, value=datetime.now().strftime("%Y-%m-%d"))
        total_bars = sum(len(inst.grids) for inst in groove_data.instruments.values())
        ws_patrones.cell(row=patron_row, column=7, value=total_bars)

        # Agregar datos a REJILLAS y HUMANIZACION
        for inst_name, inst_data in groove_data.instruments.items():
            nota_midi = INSTRUMENT_CATALOG.get(inst_name, {}).get("nota_midi", 0)

            for bar_idx, grid in enumerate(inst_data.grids):
                groove_id = self._generate_groove_id(groove_data.song_name, inst_name, bar_idx)

                # Verificar si ya existe
                if groove_id in existing_rejillas:
                    stats["skipped"] += 1
                    continue

                stats["added"] += 1

                # Agregar a REJILLAS
                row = ws_rejillas.max_row + 1
                col = 1

                ws_rejillas.cell(row=row, column=col, value=groove_id); col += 1
                ws_rejillas.cell(row=row, column=col, value=f"{inst_name}_{bar_idx + 1}"); col += 1
                ws_rejillas.cell(row=row, column=col, value=inst_name); col += 1
                ws_rejillas.cell(row=row, column=col, value=groove_data.style.value); col += 1
                ws_rejillas.cell(row=row, column=col, value=round(groove_data.bpm, 1)); col += 1
                ws_rejillas.cell(row=row, column=col, value=round(swing_amount, 1)); col += 1

                for value in grid.pattern:
                    ws_rejillas.cell(row=row, column=col, value=value); col += 1

                non_zero_vels = [v for v in grid.velocities if v > 0]
                vel_base = int(sum(non_zero_vels) / len(non_zero_vels)) if non_zero_vels else 0
                ws_rejillas.cell(row=row, column=col, value=vel_base); col += 1

                timing_var = self._calculate_timing_variance(grid.timing_deviations)
                ws_rejillas.cell(row=row, column=col, value=timing_var); col += 1

                vel_var = self._calculate_velocity_variance(grid.velocities)
                ws_rejillas.cell(row=row, column=col, value=vel_var); col += 1

                ws_rejillas.cell(row=row, column=col, value=groove_data.song_name); col += 1
                ws_rejillas.cell(row=row, column=col, value=nota_midi)

                # Agregar a HUMANIZACION
                row = ws_humanizacion.max_row + 1
                col = 1

                ws_humanizacion.cell(row=row, column=col, value=groove_id); col += 1
                ws_humanizacion.cell(row=row, column=col, value=f"{inst_name}_{bar_idx + 1}"); col += 1
                ws_humanizacion.cell(row=row, column=col, value=inst_name); col += 1

                for vel in grid.velocities:
                    ws_humanizacion.cell(row=row, column=col, value=vel); col += 1

                for dev in grid.timing_deviations:
                    ws_humanizacion.cell(row=row, column=col, value=round(dev, 2)); col += 1

                bar_duration = 4 * (60.0 / groove_data.bpm)
                ws_humanizacion.cell(row=row, column=col, value=round(bar_duration, 3)); col += 1
                ws_humanizacion.cell(row=row, column=col, value=groove_data.song_name)

        # Guardar
        wb.save(str(self.database_path))

        return stats

    def add_groove_file(self, groove_xlsx_path: str) -> dict:
        """
        Agrega datos de un archivo _groove.xlsx a database.xlsx.

        Args:
            groove_xlsx_path: Ruta al archivo _groove.xlsx

        Returns:
            Dict con estadisticas de la operacion
        """
        groove_path = Path(groove_xlsx_path)
        if not groove_path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {groove_xlsx_path}")

        # Leer el archivo groove
        wb_source = openpyxl.load_workbook(str(groove_path))

        # Extraer datos de INFO
        ws_info = wb_source["INFO"]
        song_name = None
        bpm = 75.0
        style = "one_drop"
        swing_pct = 50.0

        for row in range(3, 15):
            label = ws_info.cell(row=row, column=1).value
            value = ws_info.cell(row=row, column=2).value
            if label == "Cancion":
                song_name = value
            elif label == "BPM":
                bpm = float(value)
            elif label == "Estilo":
                style = value
            elif label == "Swing %":
                swing_pct = float(value.replace("%", ""))

        if not song_name:
            song_name = groove_path.stem.replace("_groove", "")

        # Crear GrooveData desde el archivo
        from ..models import OnsetList

        groove_data = GrooveData(
            song_name=song_name,
            bpm=bpm,
            style=JamaicanStyle(style) if style in [s.value for s in JamaicanStyle] else JamaicanStyle.ONE_DROP
        )

        if swing_pct != 50.0:
            groove_data.swing = SwingAnalysis(
                swing_percentage=swing_pct,
                swing_ratio=1.0 + (swing_pct - 50) / 50,
                is_consistent=True,
                description="Imported from groove file"
            )

        # Extraer datos de REJILLAS
        ws_rejillas = wb_source["REJILLAS"]
        ws_humanizacion = wb_source["HUMANIZACION"]

        instruments_data = {}

        for row in range(2, ws_rejillas.max_row + 1):
            id_patron = ws_rejillas.cell(row=row, column=1).value
            if not id_patron:
                continue

            inst_name = ws_rejillas.cell(row=row, column=2).value

            # Leer patron (columnas 3-18)
            pattern = []
            for col in range(3, 19):
                val = ws_rejillas.cell(row=row, column=col).value
                pattern.append(int(val) if val else 0)

            # Buscar velocidades y timing en HUMANIZACION
            velocities = [0] * 16
            timing_devs = [0.0] * 16

            for h_row in range(2, ws_humanizacion.max_row + 1):
                h_id = ws_humanizacion.cell(row=h_row, column=1).value
                if h_id == id_patron:
                    for i in range(16):
                        v = ws_humanizacion.cell(row=h_row, column=3 + i).value
                        velocities[i] = int(v) if v else 0
                        t = ws_humanizacion.cell(row=h_row, column=19 + i).value
                        timing_devs[i] = float(t) if t else 0.0
                    break

            # Agregar al instrumento
            if inst_name not in instruments_data:
                instruments_data[inst_name] = []

            instruments_data[inst_name].append(GridMapping(
                pattern=pattern,
                velocities=velocities,
                timing_deviations=timing_devs
            ))

        # Crear InstrumentData
        for inst_name, grids in instruments_data.items():
            onsets = OnsetList(onsets=[], instrument=inst_name)
            inst_data = InstrumentData(name=inst_name, onsets=onsets, grids=grids)
            groove_data.add_instrument(inst_data)

        # Agregar a database
        return self.add_groove_data(groove_data)

    def aggregate_folder(self, folder_path: str, pattern: str = "*_groove.xlsx") -> dict:
        """
        Procesa todos los archivos groove de una carpeta.

        Args:
            folder_path: Ruta a la carpeta
            pattern: Patron de archivos a buscar

        Returns:
            Dict con estadisticas totales
        """
        import glob

        folder = Path(folder_path)
        files = list(folder.glob(pattern))

        total_stats = {"files_processed": 0, "total_added": 0, "total_skipped": 0, "errors": []}

        for file_path in files:
            try:
                stats = self.add_groove_file(str(file_path))
                total_stats["files_processed"] += 1
                total_stats["total_added"] += stats["added"]
                total_stats["total_skipped"] += stats["skipped"]
                print(f"  Procesado: {file_path.name} (+{stats['added']} patrones)")
            except Exception as e:
                total_stats["errors"].append({"file": str(file_path), "error": str(e)})
                print(f"  Error en {file_path.name}: {e}")

        return total_stats

    def create_empty_database(self) -> str:
        """
        Crea un database.xlsx vacio con la estructura base.

        Returns:
            Ruta del archivo creado
        """
        wb = self._create_base_database()
        wb.save(str(self.database_path))
        return str(self.database_path)

    def get_statistics(self) -> dict:
        """
        Obtiene estadisticas de la base de datos.

        Returns:
            Dict con estadisticas
        """
        if not self.database_path.exists():
            return {"error": "Database no existe"}

        wb = openpyxl.load_workbook(str(self.database_path))

        stats = {
            "estilos": wb["ESTILOS"].max_row - 1,
            "patrones": wb["PATRONES"].max_row - 1,
            "rejillas": wb["REJILLAS"].max_row - 1,
            "humanizacion": wb["HUMANIZACION"].max_row - 1,
            "instrumentos": wb["INSTRUMENTOS"].max_row - 1,
        }

        # Contar por estilo
        ws_rejillas = wb["REJILLAS"]
        estilos_count = {}
        for row in range(2, ws_rejillas.max_row + 1):
            estilo = ws_rejillas.cell(row=row, column=4).value
            if estilo:
                estilos_count[estilo] = estilos_count.get(estilo, 0) + 1

        stats["por_estilo"] = estilos_count

        return stats
